#!/usr/bin/env python3
"""
BKNS Wiki — Excel Direct Injection (Ground Truth)

Parses the official BKNS pricing Excel file and generates
ground_truth claims bypassing LLM entirely.

This ensures 100% accurate pricing data in the wiki.

Usage:
    python3 tools/enrich_excel.py                     # Parse all sheets
    python3 tools/enrich_excel.py --sheet VPS          # Parse specific sheet
    python3 tools/enrich_excel.py --dry-run             # Preview without writing
"""
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime, timezone, timedelta

import openpyxl
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from lib.config import CLAIMS_APPROVED_DIR, CLAIMS_DRAFTS_DIR
from lib.utils import ensure_dir
from lib.logger import log_entry

# ── Constants ──────────────────────────────────────────────
EXCEL_PATH = Path("/home/openclaw/wiki/raw/manual/Bảng giá Hosting- VPS- Email.xlsx")
SOURCE_ID = "EXCEL-BANGGIA-BKNS-2026"
VN_TZ = timezone(timedelta(hours=7))
NOW_ISO = datetime.now(VN_TZ).isoformat()
TODAY = datetime.now(VN_TZ).strftime("%Y-%m-%d")


def _safe_id(text: str) -> str:
    """Normalize text to safe ID: lowercase, ascii-ish, underscore-separated."""
    text = str(text).lower().strip()
    text = text.replace("+", "_plus")
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')


def _claim_filename(claim_id: str) -> str:
    """Generate safe filename from claim_id."""
    safe = _safe_id(claim_id)
    # Remove leading "clm_" if already present from CLM- prefix
    if safe.startswith("clm_"):
        safe = safe[4:]
    return f"clm_{safe}.yaml"


def _write_claim(claim: dict, category_dir: str, dry_run: bool = False) -> str:
    """Write a single claim YAML file. Returns the filepath."""
    out_dir = CLAIMS_APPROVED_DIR / category_dir
    ensure_dir(out_dir)
    
    fname = _claim_filename(claim["claim_id"])
    fpath = out_dir / fname
    
    if dry_run:
        return str(fpath)
    
    with open(fpath, "w", encoding="utf-8") as f:
        yaml.dump(claim, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    
    return str(fpath)


def _make_claim(
    entity_id: str,
    entity_name: str,
    attribute: str,
    value,
    unit: str = "",
    qualifiers: dict = None,
    compiler_note: str = "",
    entity_type: str = "product_plan",
) -> dict:
    """Build a standardized claim dict with ground_truth confidence."""
    claim_id = f"CLM-EXCEL-{_safe_id(entity_id)}-{_safe_id(attribute)}"
    
    claim = {
        "claim_id": claim_id,
        "entity_id": entity_id,
        "entity_type": entity_type,
        "entity_name": entity_name,
        "attribute": attribute,
        "value": value,
        "confidence": "ground_truth",
        "review_state": "approved",
        "risk_class": "verified",
        "source_ids": [SOURCE_ID],
        "observed_at": NOW_ISO,
        "valid_from": TODAY,
        "approved_at": NOW_ISO,
        "approved_by": "excel-direct-injection",
    }
    
    if unit:
        claim["unit"] = unit
    if qualifiers:
        claim["qualifiers"] = qualifiers
    if compiler_note:
        claim["compiler_note"] = compiler_note
    
    return claim


# ── VPS Parser ─────────────────────────────────────────────
def parse_vps(ws) -> list[dict]:
    """Parse VPS sheet into claims."""
    claims = []
    
    # Read all rows into a matrix for easier access
    rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_num = row[0].row
        cells = {}
        for c in row:
            if c.value is not None:
                cells[c.column] = c.value
        if cells:
            rows[row_num] = cells
    
    # Detect product sections by finding "Thuộc tính" rows
    sections = []
    for r, cells in sorted(rows.items()):
        val = str(cells.get(1, ""))
        # Detect section headers like "1. VPS ADM", "2. VPS VM"
        if re.match(r'^\d+\.\s+', val) or val == "Bảng mới":
            sections.append({"header_row": r, "name": val})
    
    # Parse each section
    for idx, section in enumerate(sections):
        start = section["header_row"]
        end = sections[idx + 1]["header_row"] - 1 if idx + 1 < len(sections) else ws.max_row
        
        section_name = section["name"]
        
        # Determine product line
        if "ADM" in section_name.upper():
            product_line = "vps-amd"
            display_prefix = "Cloud VPS AMD"
        elif "VM" in section_name.upper():
            product_line = "vps-vm"
            display_prefix = "Cloud VPS VM"
        elif "MMO" in section_name.upper() or "Giá rẻ" in section_name:
            product_line = "vps-mmo"
            display_prefix = "VPS MMO/Giá rẻ"
        elif "Bảng mới" in section_name:
            product_line = "vps-mmo-new"
            display_prefix = "VPS MMO (Bảng mới)"
        else:
            product_line = _safe_id(section_name)
            display_prefix = section_name
        
        # Find the "Thuộc tính" row to get plan codes
        plan_codes = {}
        spec_rows = {}
        price_rows = {}
        
        for r in range(start, end + 1):
            if r not in rows:
                continue
            cells = rows[r]
            label = str(cells.get(1, cells.get(2, ""))).strip().replace("\n", " ")
            
            if "Thuộc tính" in label:
                for col, val in cells.items():
                    if col >= 2 and col <= 9 and val and "bổ sung" not in str(val):
                        plan_codes[col] = str(val).strip()
            elif any(kw in label for kw in ["CPU", "RAM", "SSD", "Dung lượng", "NVMe", "Tốc độ", "IPv4", "IPv6"]):
                spec_rows[label] = {col: val for col, val in cells.items() if col >= 2 and col <= 9}
            elif "Giá" in label or "tháng" in label:
                price_rows[label] = {col: val for col, val in cells.items() if col >= 2 and col <= 9}
        
        if not plan_codes:
            continue
        
        # Generate claims for each plan
        for col, plan_code in plan_codes.items():
            entity_id = f"product.vps.{_safe_id(plan_code)}"
            entity_name = f"{display_prefix} {plan_code}"
            
            # Spec claims
            for spec_label, spec_vals in spec_rows.items():
                val = spec_vals.get(col)
                if val is None:
                    continue
                
                # Determine attribute
                clean_label = spec_label.replace("\n", " ").strip()
                if "CPU" in clean_label:
                    attr = "cpu_cores"
                elif "RAM" in clean_label:
                    attr = "ram"
                elif "SSD" in clean_label or "Dung lượng" in clean_label or "NVMe" in clean_label:
                    attr = "storage"
                elif "Tốc độ" in clean_label:
                    attr = "network_speed"
                elif "IPv4" in clean_label:
                    attr = "ipv4"
                elif "IPv6" in clean_label:
                    attr = "ipv6"
                else:
                    attr = _safe_id(clean_label)
                
                claims.append(_make_claim(
                    entity_id=entity_id,
                    entity_name=entity_name,
                    attribute=attr,
                    value=str(val),
                    compiler_note=f"Excel VPS sheet, {section_name}, plan {plan_code}, {clean_label}",
                ))
            
            # Price claims  
            for price_label, price_vals in price_rows.items():
                val = price_vals.get(col)
                if val is None:
                    continue
                
                clean_label = price_label.replace("\n", " ").strip()
                
                # Determine price period
                if "1 tháng" in clean_label or "gốc" in clean_label or "trước giảm" in clean_label:
                    attr = "monthly_price"
                    period = "1 tháng"
                elif "3 tháng" in clean_label:
                    attr = "price_3m"
                    period = "3 tháng"
                elif "6" in clean_label and "tháng" in clean_label.lower():
                    attr = "price_6m"
                    period = "6 tháng"
                elif "12 tháng" in clean_label:
                    attr = "price_12m"
                    period = "12 tháng"
                elif "24 tháng" in clean_label:
                    attr = "price_24m"
                    period = "24 tháng"
                elif "36 tháng" in clean_label:
                    attr = "price_36m"
                    period = "36 tháng"
                elif "60 tháng" in clean_label:
                    attr = "price_60m"
                    period = "60 tháng"
                else:
                    attr = f"price_{_safe_id(clean_label)}"
                    period = clean_label
                
                # Extract discount if present
                discount_match = re.search(r'\(?\s*-?\s*(\d+)%\s*\)?', clean_label)
                qualifiers = {"period": period, "vat_included": False}
                if discount_match:
                    qualifiers["discount_pct"] = int(discount_match.group(1))
                
                claims.append(_make_claim(
                    entity_id=entity_id,
                    entity_name=entity_name,
                    attribute=attr,
                    value=int(val) if isinstance(val, (int, float)) else val,
                    unit="VND",
                    qualifiers=qualifiers,
                    compiler_note=f"Excel VPS sheet, {section_name}, plan {plan_code}, {clean_label}",
                ))
    
    return claims


# ── Hosting Parser ─────────────────────────────────────────
def parse_hosting(ws) -> list[dict]:
    """Parse Hosting sheet into claims."""
    claims = []
    
    rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_num = row[0].row
        cells = {}
        for c in row:
            if c.value is not None:
                cells[c.column] = c.value
        if cells:
            rows[row_num] = cells
    
    sections = []
    for r, cells in sorted(rows.items()):
        val = str(cells.get(1, ""))
        if re.match(r'^\d+\.\s+', val):
            sections.append({"header_row": r, "name": val})
    
    for idx, section in enumerate(sections):
        start = section["header_row"]
        end = sections[idx + 1]["header_row"] - 1 if idx + 1 < len(sections) else ws.max_row
        section_name = section["name"]
        
        # Determine hosting type
        if "Platium" in section_name or "platium" in section_name.lower():
            product_line = "hosting-linux"
            display_prefix = "Platium Web Hosting"
        elif "Wordpress" in section_name:
            product_line = "hosting-wordpress"
            display_prefix = "Hosting WordPress"
        elif "Windows" in section_name:
            product_line = "hosting-windows"
            display_prefix = "Hosting Windows"
        else:
            product_line = _safe_id(section_name)
            display_prefix = section_name
        
        plan_codes = {}
        spec_rows = {}
        price_rows = {}
        
        for r in range(start, end + 1):
            if r not in rows:
                continue
            cells = rows[r]
            label = str(cells.get(1, "")).strip().replace("\n", " ")
            
            if "Thuộc tính" in label:
                for col, val in cells.items():
                    if col >= 2 and col <= 9 and val:
                        plan_codes[col] = str(val).strip()
            elif any(kw in label for kw in ["CPU", "RAM", "NVMe", "SSD", "Disk", "Domain"]):
                spec_rows[label] = {col: val for col, val in cells.items() if col >= 2 and col <= 9}
            elif "Giá" in label:
                price_rows[label] = {col: val for col, val in cells.items() if col >= 2 and col <= 9}
        
        if not plan_codes:
            continue
        
        for col, plan_code in plan_codes.items():
            entity_id = f"product.hosting.{_safe_id(plan_code)}"
            entity_name = f"{display_prefix} {plan_code}"
            
            # Spec claims
            for spec_label, spec_vals in spec_rows.items():
                val = spec_vals.get(col)
                if val is None:
                    continue
                
                clean_label = spec_label.replace("\n", " ").strip()
                if "CPU" in clean_label:
                    attr = "cpu_cores"
                elif "RAM" in clean_label:
                    attr = "ram"
                elif "NVMe" in clean_label or "SSD" in clean_label or "Disk" in clean_label:
                    attr = "storage"
                elif "Domain" in clean_label:
                    attr = "domain_count"
                else:
                    attr = _safe_id(clean_label)
                
                claims.append(_make_claim(
                    entity_id=entity_id,
                    entity_name=entity_name,
                    attribute=attr,
                    value=str(val) if not isinstance(val, (int, float)) else val,
                    compiler_note=f"Excel Hosting sheet, {section_name}, plan {plan_code}, {clean_label}",
                ))
            
            # Price claims
            for price_label, price_vals in price_rows.items():
                val = price_vals.get(col)
                if val is None:
                    continue
                
                clean_label = price_label.replace("\n", " ").strip()
                
                if "tháng" in clean_label and "năm" not in clean_label:
                    attr = "monthly_price"
                    period = "1 tháng"
                elif "1 năm" in clean_label:
                    attr = "price_1y"
                    period = "1 năm"
                elif "2 năm" in clean_label:
                    attr = "price_2y"
                    period = "2 năm"
                elif "3 năm" in clean_label:
                    attr = "price_3y"
                    period = "3 năm"
                elif "4 năm" in clean_label:
                    attr = "price_4y"
                    period = "4 năm"
                else:
                    attr = f"price_{_safe_id(clean_label)}"
                    period = clean_label
                
                discount_match = re.search(r'\(?\s*-?\s*(\d+)%\s*\)?', clean_label)
                qualifiers = {"period": period, "vat_included": False}
                if discount_match:
                    qualifiers["discount_pct"] = int(discount_match.group(1))
                
                claims.append(_make_claim(
                    entity_id=entity_id,
                    entity_name=entity_name,
                    attribute=attr,
                    value=int(val) if isinstance(val, (int, float)) else val,
                    unit="VND",
                    qualifiers=qualifiers,
                    compiler_note=f"Excel Hosting sheet, {section_name}, plan {plan_code}, {clean_label}",
                ))
    
    return claims


# ── Email Parser ───────────────────────────────────────────
def parse_email(ws) -> list[dict]:
    """Parse Email sheet into claims."""
    claims = []
    
    rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_num = row[0].row
        cells = {}
        for c in row:
            if c.value is not None:
                cells[c.column] = c.value
        if cells:
            rows[row_num] = cells
    
    # ── Email Hosting (rows ~3-13) ──
    # Header row has multi-line plan descriptions
    header_row = rows.get(4, {})
    email_hosting_plans = {}
    for col in [2, 3, 4, 5]:
        desc = header_row.get(col)
        if desc:
            # Parse: "EMAIL 1\n42.750đ/ tháng\nSố lượng Email: 05..."
            lines = str(desc).split("\n")
            plan_name = lines[0].strip()
            plan_code = _safe_id(plan_name)
            
            # Extract specs from description
            specs = {}
            for line in lines[1:]:
                if "Số lượng Email" in line:
                    match = re.search(r':\s*(\d+)', line)
                    if match:
                        specs["email_accounts"] = int(match.group(1))
                elif "Forwarders" in line:
                    match = re.search(r':\s*(\d+)', line)
                    if match:
                        specs["forwarders"] = int(match.group(1))
                elif "Dung lượng" in line:
                    match = re.search(r':\s*(.+)', line)
                    if match:
                        specs["storage_per_email"] = match.group(1).strip()
            
            email_hosting_plans[col] = {
                "name": plan_name,
                "code": plan_code,
                "specs": specs,
            }
    
    # Generate spec claims for Email Hosting
    for col, plan_info in email_hosting_plans.items():
        entity_id = f"product.email.hosting.{plan_info['code']}"
        entity_name = f"Email Hosting {plan_info['name']}"
        
        for spec_attr, spec_val in plan_info["specs"].items():
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute=spec_attr,
                value=spec_val,
                compiler_note=f"Excel Email sheet, Email Hosting, {plan_info['name']}",
            ))
    
    # Price rows for Email Hosting (rows 5-9)
    period_map = {5: ("monthly_price", "1 tháng"), 6: ("price_1y", "1 năm"),
                  7: ("price_2y", "2 năm"), 8: ("price_3y", "3 năm"),
                  9: ("price_4y", "4 năm")}
    
    for row_num, (attr, period) in period_map.items():
        if row_num not in rows:
            continue
        cells = rows[row_num]
        for col, plan_info in email_hosting_plans.items():
            val = cells.get(col)
            if val is None or not isinstance(val, (int, float)):
                continue
            entity_id = f"product.email.hosting.{plan_info['code']}"
            entity_name = f"Email Hosting {plan_info['name']}"
            
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute=attr,
                value=int(val),
                unit="VND",
                qualifiers={"period": period, "vat_included": False},
                compiler_note=f"Excel Email sheet, Email Hosting, {plan_info['name']}, {period}",
            ))
    
    # ── Email Server - Mini (rows ~16-21) ──
    mini_header = rows.get(16, {})
    mini_plans = {}
    for col in [2, 3, 4, 5]:
        desc = mini_header.get(col)
        if desc:
            lines = str(desc).split("\n")
            plan_name = lines[0].strip()
            storage = ""
            for line in lines[1:]:
                if "Dung lượng" in line:
                    match = re.search(r':\s*(.+)', line)
                    if match:
                        storage = match.group(1).strip()
            
            mini_plans[col] = {"name": plan_name, "code": _safe_id(plan_name), "storage": storage}
    
    for col, plan_info in mini_plans.items():
        entity_id = f"product.email.server.{plan_info['code']}"
        entity_name = f"Cloud Email Server {plan_info['name']}"
        
        if plan_info["storage"]:
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute="storage",
                value=plan_info["storage"],
                compiler_note=f"Excel Email sheet, Email Server Mini, {plan_info['name']}",
            ))
    
    mini_period_map = {17: ("monthly_price", "1 tháng"), 18: ("price_6m", "6 tháng"),
                       19: ("price_1y", "1 năm"), 20: ("price_2y", "2 năm"),
                       21: ("price_3y", "3 năm")}
    
    for row_num, (attr, period) in mini_period_map.items():
        if row_num not in rows:
            continue
        cells = rows[row_num]
        for col, plan_info in mini_plans.items():
            val = cells.get(col)
            if val is None or not isinstance(val, (int, float)):
                continue
            entity_id = f"product.email.server.{plan_info['code']}"
            entity_name = f"Cloud Email Server {plan_info['name']}"
            
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute=attr,
                value=int(val),
                unit="VND",
                qualifiers={"period": period, "vat_included": False},
                compiler_note=f"Excel Email sheet, Email Server, {plan_info['name']}, {period}",
            ))
    
    # ── Email Server - ES (rows ~28-33) ──
    es_header = rows.get(28, {})
    es_plans = {}
    for col in [2, 3, 4, 5]:
        desc = es_header.get(col)
        if desc:
            lines = str(desc).split("\n")
            plan_name = lines[0].strip()
            storage = ""
            for line in lines[1:]:
                if "Dung lượng" in line:
                    match = re.search(r':\s*(.+)', line)
                    if match:
                        storage = match.group(1).strip()
            es_plans[col] = {"name": plan_name, "code": _safe_id(plan_name), "storage": storage}
    
    for col, plan_info in es_plans.items():
        entity_id = f"product.email.server.{plan_info['code']}"
        entity_name = f"Cloud Email Server {plan_info['name']}"
        
        if plan_info["storage"]:
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute="storage",
                value=plan_info["storage"],
                compiler_note=f"Excel Email sheet, Email Server ES, {plan_info['name']}",
            ))
    
    es_period_map = {29: ("monthly_price", "1 tháng"), 30: ("price_6m", "6 tháng"),
                     31: ("price_1y", "1 năm"), 32: ("price_2y", "2 năm"),
                     33: ("price_3y", "3 năm")}
    
    for row_num, (attr, period) in es_period_map.items():
        if row_num not in rows:
            continue
        cells = rows[row_num]
        for col, plan_info in es_plans.items():
            val = cells.get(col)
            if val is None or not isinstance(val, (int, float)):
                continue
            entity_id = f"product.email.server.{plan_info['code']}"
            entity_name = f"Cloud Email Server {plan_info['name']}"
            
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute=attr,
                value=int(val),
                unit="VND",
                qualifiers={"period": period, "vat_included": False},
                compiler_note=f"Excel Email sheet, Email Server ES, {plan_info['name']}, {period}",
            ))
    
    # ── Email Relay (rows ~39-45) ──
    relay_header = rows.get(40, {})
    relay_plans = {}
    for col in [3, 4, 5, 6]:
        val = relay_header.get(col)
        if val:
            relay_plans[col] = {"name": str(val).strip(), "code": _safe_id(str(val))}
    
    # Relay specs (row 41 = mail/ngày)
    if 41 in rows:
        for col, plan_info in relay_plans.items():
            val = rows[41].get(col)
            if val:
                entity_id = f"product.email.relay.{plan_info['code']}"
                entity_name = f"Email Relay {plan_info['name']}"
                claims.append(_make_claim(
                    entity_id=entity_id,
                    entity_name=entity_name,
                    attribute="daily_email_limit",
                    value=int(val) if isinstance(val, (int, float)) else val,
                    unit="emails/ngày",
                    compiler_note=f"Excel Email sheet, Email Relay, {plan_info['name']}",
                ))
    
    relay_period_map = {42: ("monthly_price", "1 tháng"), 43: ("price_1y", "1 năm"),
                        44: ("price_2y", "2 năm"), 45: ("price_3y", "3 năm")}
    
    for row_num, (attr, period) in relay_period_map.items():
        if row_num not in rows:
            continue
        cells = rows[row_num]
        for col, plan_info in relay_plans.items():
            val = cells.get(col)
            if val is None or not isinstance(val, (int, float)):
                continue
            entity_id = f"product.email.relay.{plan_info['code']}"
            entity_name = f"Email Relay {plan_info['name']}"
            
            claims.append(_make_claim(
                entity_id=entity_id,
                entity_name=entity_name,
                attribute=attr,
                value=int(val),
                unit="VND",
                qualifiers={"period": period, "vat_included": False},
                compiler_note=f"Excel Email sheet, Email Relay, {plan_info['name']}, {period}",
            ))
    
    return claims


# ── Category router ────────────────────────────────────────
CATEGORY_ROUTES = {
    "product.vps": "products/vps",
    "product.hosting": "products/hosting",
    "product.email": "products/email",
}


def route_category(entity_id: str) -> str:
    """Route entity_id to category directory."""
    for prefix, cat_dir in CATEGORY_ROUTES.items():
        if entity_id.startswith(prefix):
            return cat_dir
    return "products/other"


# ── Main ───────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="BKNS Excel → Ground Truth Claims")
    parser.add_argument("--sheet", choices=["VPS", "Hosting", "Email", "all"], default="all",
                        help="Which sheet to parse")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't write files")
    args = parser.parse_args()
    
    if not EXCEL_PATH.exists():
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    
    all_claims = []
    parsers = {
        "VPS": parse_vps,
        "Hosting": parse_hosting,
        "Email": parse_email,
    }
    
    sheets_to_parse = list(parsers.keys()) if args.sheet == "all" else [args.sheet]
    
    for sheet_name in sheets_to_parse:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found in Excel")
            continue
        
        ws = wb[sheet_name]
        sheet_claims = parsers[sheet_name](ws)
        all_claims.extend(sheet_claims)
        print(f"📊 {sheet_name}: {len(sheet_claims)} claims extracted")
    
    # Write claims
    written = 0
    skipped = 0
    by_category = {}
    
    for claim in all_claims:
        cat_dir = route_category(claim["entity_id"])
        by_category.setdefault(cat_dir, 0)
        
        fpath = _write_claim(claim, cat_dir, dry_run=args.dry_run)
        by_category[cat_dir] += 1
        written += 1
    
    # Summary
    print(f"\n{'═' * 50}")
    print(f"📋 Total claims: {len(all_claims)}")
    print(f"✅ Written: {written}" + (" (dry-run)" if args.dry_run else ""))
    print(f"\nBy category:")
    for cat, count in sorted(by_category.items()):
        print(f"  {cat}: {count}")
    
    if not args.dry_run:
        log_entry("enrich-excel", "success",
                  f"Injected {written} ground_truth claims from Excel",
                  extra={"by_category": by_category})
        print(f"\n🎯 Ground truth claims injected to: {CLAIMS_APPROVED_DIR}/")


if __name__ == "__main__":
    main()
